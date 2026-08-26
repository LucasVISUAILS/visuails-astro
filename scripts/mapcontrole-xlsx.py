# -*- coding: utf-8 -*-
"""
VISUAILS — de bevindingen van de mapcontrole als werkblad.
    python3 scripts/mapcontrole-xlsx.py [uitvoerpad]

Eén rij per wijziging die doorgevoerd moet worden. Twee tabbladen: "Lees mij"
met de uitleg en één ingevulde voorbeeldrij, en "Wijzigingen" met alles.

Alleen de twee gele kolommen worden ingevuld. Er staan geen formules in dit
bestand, dus er valt niets te herberekenen.
"""
import io, json, sys, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

rijen = json.load(io.open('/tmp/wijzigingen.json', encoding='utf-8'))
ORDE = {'breekt': 0, 'vraag': 1, 'juridisch': 2, 'verouderd': 3}
rijen.sort(key=lambda r: (0 if r['actie'].startswith('al gedaan') else 1, ORDE.get(r['ernst'], 9), r['nr']))

INKT, GROEN, KOPBALK, LICHT, INVUL = '14150F', '5E6E00', '1C1E14', 'F4F4EC', 'FBF6D8'
ROOD, AMBER, BLAUW, LIJN = '9C2A1B', '8A5008', '2A557A', 'D8D8CC'

wb = Workbook()

# ══ Lees mij ════════════════════════════════════════════════════════════════
lm = wb.active; lm.title = 'Lees mij'; lm.sheet_properties.tabColor = GROEN
lm.column_dimensions['A'].width = 24
lm.column_dimensions['B'].width = 104

def zet(cel, waarde, vet=False, maat=11, kleur=INKT, wrap=False, cursief=False):
    c = lm[cel]; c.value = waarde
    c.font = Font(name='Arial', size=maat, bold=vet, italic=cursief, color=kleur)
    c.alignment = Alignment(wrap_text=wrap, vertical='top')
    return c

zet('A1', 'VISUAILS · controle van de projectmap', vet=True, maat=18)
zet('A2', datetime.date.today().strftime('Gegenereerd op %d-%m-%Y uit de map en de code'), maat=10, kleur='6B6D5C')

rij = 4
def blok(kop, regels):
    global rij
    zet(f'A{rij}', kop, vet=True, maat=12, kleur=GROEN); rij += 1
    for r in regels:
        zet(f'B{rij}', r, wrap=True)
        lm.row_dimensions[rij].height = 15 * (1 + len(r) // 96)
        rij += 1
    rij += 1

blok('Wat dit is', [
 f'{len(rijen)} plekken in de projectmap waar de tekst iets beweert dat niet meer klopt. Elke rij is één '
 'wijziging: het bestand, de regel, wat er nu staat, en wat het moet worden.',
 'Een document dat de geschiedenis beschrijft ("tot 14 augustus stond hier X") staat hier NIET in. Dat is '
 'een verslag en dat hoort te blijven staan. Wat hier staat, zegt in de tegenwoordige tijd iets onwaars.',
])

blok('Wat jij doet', [
 'Lees de kolommen "staat er nu" en "wordt". Ben je het ermee eens, zet dan "ja" in de kolom akkoord — '
 'dan voer ik het door.',
 'Niet mee eens, of wil je het anders? Zet "nee" of "anders" en schrijf in de laatste kolom wat er in '
 'plaats daarvan moet komen.',
 'Laat je een rij leeg, dan verandert er niets aan dat bestand.',
 'Stuur het bestand daarna terug in Cowork. Ik lees per rij; het bestand en de regel staan erbij, dus er '
 'kan niets door elkaar lopen.',
])

blok('De kolom "actie"', [
 'al gedaan — dit heb ik al gerepareerd en het staat op je schijf. Hier hoef je niets mee, behalve waar '
 'erbij staat dat je de formulering naleest.',
 'vervangen — de regel in "wordt" gaat één op één de plaats innemen van wat er nu staat.',
 'herschrijven — de fout is te groot voor één regel; "wordt" beschrijft wat er moet gebeuren.',
 'schrappen — de passage beschrijft iets dat niet meer bestaat en kan weg.',
 'afvinken — een openstaand punt in de werklijst dat af is.',
 'jij controleert / jij kiest — hier kan ik het niet voor je beslissen.',
])

blok('De kolom "ernst"', [
 'breekt — wie dit document vandaag volgt, maakt iets stuk of raakt tijd kwijt. Bovenaan.',
 'juridisch — het verwerkingsregister en de datalekprocedure. Die zijn niet intern: art. 30 AVG vraagt dat '
 'het register de werkelijkheid beschrijft.',
 'verouderd — beschrijft het project van vóór augustus. Niet dringend, maar wie het leest om iets te weten '
 'te komen, komt verkeerd uit.',
])

zet(f'A{rij}', 'Zo ziet een ingevulde rij eruit', vet=True, maat=12, kleur=GROEN); rij += 2
vb = [('bestand', 'README.md'), ('regel', 'r3'), ('staat er nu', "92 pagina's"),
      ('wordt', "90 pagina's"), ('akkoord?', 'ja'), ('jouw opmerking', '')]
for k, (kop, waarde) in enumerate(vb):
    c = lm.cell(row=rij, column=k + 1, value=kop)
    c.font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor=KOPBALK)
    c.alignment = Alignment(wrap_text=True, vertical='center')
    d = lm.cell(row=rij + 1, column=k + 1, value=waarde)
    d.font = Font(name='Arial', size=10, color=INKT)
    d.alignment = Alignment(wrap_text=True, vertical='top')
    if kop in ('akkoord?', 'jouw opmerking'):
        d.fill = PatternFill('solid', fgColor=INVUL)
for k, b in enumerate([24, 104, 26, 30, 14, 30]):
    lm.column_dimensions[get_column_letter(k + 1)].width = b
rij += 3
zet(f'A{rij}', 'Alleen de twee gele kolommen vul je in.', cursief=True, maat=10, kleur=AMBER)

# ══ Wijzigingen ═════════════════════════════════════════════════════════════
ws = wb.create_sheet('Wijzigingen'); ws.sheet_properties.tabColor = KOPBALK
KOL = [('nr', 6), ('ernst', 12), ('actie', 20), ('bestand', 26), ('regel', 20),
       ('staat er nu', 56), ('wordt', 62), ('waarom', 56), ('akkoord?', 13), ('jouw opmerking', 34)]
for i, (naam, breedte) in enumerate(KOL, start=1):
    c = ws.cell(row=1, column=i, value=naam)
    c.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor=KOPBALK)
    c.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')
    ws.column_dimensions[get_column_letter(i)].width = breedte
ws.row_dimensions[1].height = 26

KLEUR = {'breekt': ROOD, 'juridisch': AMBER, 'vraag': BLAUW, 'verouderd': '6B6D5C'}
grens = Border(top=Side(style='thin', color=LIJN))
vorige = None
for n, r in enumerate(rijen, start=2):
    gedaan = r['actie'].startswith('al gedaan')
    waarden = [r['nr'], r['ernst'], r['actie'], r['bestand'], r['regel'],
               r['nu'], r['wordt'], r['waarom'], '', '']
    nieuw = r['ernst'] != vorige; vorige = r['ernst']
    for k, v in enumerate(waarden, start=1):
        c = ws.cell(row=n, column=k, value=v)
        c.font = Font(name='Arial', size=10, bold=(k == 2),
                      color=KLEUR.get(r['ernst'], INKT) if k == 2 else ('6B6D5C' if k in (1, 5, 8) else INKT))
        c.alignment = Alignment(wrap_text=k >= 4, vertical='top')
        if k in (9, 10) and not gedaan:
            c.fill = PatternFill('solid', fgColor=INVUL)
        elif k <= 3 or gedaan:
            c.fill = PatternFill('solid', fgColor=LICHT)
        if nieuw:
            c.border = grens

ws.freeze_panes = 'F2'
ws.auto_filter.ref = f'A1:J{len(rijen) + 1}'

pad = sys.argv[1] if len(sys.argv) > 1 else 'tekstronde/VISUAILS-mapcontrole.xlsx'
wb.save(pad)
print('geschreven:', pad, '·', len(rijen), 'rijen')
