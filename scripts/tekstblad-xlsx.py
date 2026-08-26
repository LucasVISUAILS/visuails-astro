# -*- coding: utf-8 -*-
"""
VISUAILS — tekstronde/blad.json omzetten naar één werkblad om na te kijken.
    npm run build && npm run tekstblad && python3 scripts/tekstblad-xlsx.py

Twee tabbladen: "Lees mij" met de uitleg en één ingevulde voorbeeldrij, en
"Tekst" met alle regels. Alleen de twee gele kolommen worden ingevuld; alle
andere kolommen zijn hoe een teruggestuurde regel herkend wordt.

Er staan geen formules in dit bestand, dus er valt niets te herberekenen — het
is een tabel om te lezen en in te vullen, geen model.
"""
import json, io, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

rijen = json.load(io.open('tekstronde/blad.json', encoding='utf-8'))

PAGINA_ORDE = ['home','catalog','lifestyle','video','merkmodel','prijzen','abonnementen',
 'hoe-het-werkt','vergelijken','studio','modellen','galerij','over-ons','faq','gidsen',
 'uploadrichtlijnen','proefvisual','start','start-catalog','start-lifestyle','start-allebei',
 'start-video','start-merkmodel','start-eigen-look','start-abonnement','contact','bedankt',
 'portaal','voorwaarden','privacy','cookies','ai-act','verwerkersovereenkomst']
GROEP_ORDE = ['Pagina', 'Menu en voettekst', 'VISUAILS Studio', 'E-mail', 'WhatsApp']

for i, r in enumerate(rijen):
    r['_i'] = i
    r['_g'] = GROEP_ORDE.index(r['groep']) if r['groep'] in GROEP_ORDE else 9
    r['_p'] = PAGINA_ORDE.index(r['slug']) if r['slug'] in PAGINA_ORDE else 99
rijen.sort(key=lambda r: (r['_g'], r['_p'], r['_i']))

# ── kleuren, afgeleid van de site zelf ──────────────────────────────────────
INKT      = '14150F'
GROEN     = '5E6E00'
KOPBALK   = '1C1E14'
LICHT     = 'F4F4EC'
INVUL     = 'FBF6D8'
LETOP     = '8A5008'
LIJN      = 'D8D8CC'

wb = Workbook()

# ══ blad 1 · Lees mij ═══════════════════════════════════════════════════════
lm = wb.active
lm.title = 'Lees mij'
lm.sheet_properties.tabColor = GROEN

def zet(cel, waarde, *, vet=False, maat=11, kleur=INKT, wrap=False, vul=None, cursief=False):
    c = lm[cel]; c.value = waarde
    c.font = Font(name='Arial', size=maat, bold=vet, italic=cursief, color=kleur)
    c.alignment = Alignment(wrap_text=wrap, vertical='top')
    if vul: c.fill = PatternFill('solid', fgColor=vul)
    return c

lm.column_dimensions['A'].width = 22
lm.column_dimensions['B'].width = 104

zet('A1', 'VISUAILS · tekstronde', vet=True, maat=18)
zet('A2', datetime.date.today().strftime('Gegenereerd op %d-%m-%Y uit de gebouwde site en de broncode'), maat=10, kleur='6B6D5C')

rij = 4
def blok(kop, regels):
    global rij
    zet(f'A{rij}', kop, vet=True, maat=12, kleur=GROEN)
    rij += 1
    for r in regels:
        zet(f'B{rij}', r, wrap=True)
        lm.row_dimensions[rij].height = 15 * (1 + len(r) // 96)
        rij += 1
    rij += 1

blok('Wat je hier hebt', [
 f'Alle {len(rijen)} zinnen die een klant van VISUAILS te lezen krijgt, op één tabblad: de 33 publieke '
 'pagina\'s in beide talen, het menu en de voettekst, de schermen in VISUAILS Studio, de e-mails die uitgaan, '
 'en de voorgevulde WhatsApp-berichten.',
 'Ze staan in de volgorde waarin je ze tegenkomt — van de voordeur naar de kassa, daarna het portaal, '
 'daarna de berichten. Niet alfabetisch.',
])

blok('Wat jij doet', [
 'Lees de kolommen Engels nu en Nederlands nu. Klopt de zin, doe je niets — leeg betekent ongewijzigd.',
 'Wil je hem anders, typ je versie in Engels nieuw en/of Nederlands nieuw. Dat zijn de twee gele kolommen '
 'en het zijn de enige die je invult.',
 'Je hoeft niet in één keer door. Filter op de kolom waar en doe één pagina, sla op, ga later verder.',
 'Stuur het bestand daarna terug in Cowork. Ik lees per rij: het adres in de kolom bron zegt exact waar '
 'die zin in de code staat, dus er kan niets door elkaar lopen.',
])

blok('Wat je NIET moet aanpassen', [
 'De kolommen id, groep, waar, blok, soort, bron en let op. Daar herken ik de rij aan.',
 'Rijen verwijderen of de volgorde omgooien hoeft niet en helpt niet — filteren doet hetzelfde en houdt '
 'het bestand heel.',
])

blok('De kolom "let op"', [
 'Staat er "de code vult in: …", dan wordt dat stuk van de zin door de code ingevuld en is het geen tekst '
 'die je uittypt. Schrijf je "€1.190" voluit, dan staat dat bedrag straks vast en beweegt het niet meer mee '
 'met de prijslijst. Laat dat stuk staan zoals het er staat, of beschrijf wat je wilt — ik zet de verwijzing '
 'er weer in.',
 'Dat is precies wat de vorige ronde 31 zinnen kostte, dus het staat er nu bij.',
 'Staat er "staat ook op N andere pagina\'s", dan verander je met deze ene rij die N pagina\'s tegelijk.',
])

blok('De kolom "bron", zodat je hem kunt lezen', [
 'HomeV2.astro › en.priceLede — het bestand, en daarin de sleutel priceLede in de Engelse helft.',
 'HomeV2.astro › en.svcH[1] — dezelfde sleutel, maar de tweede tekst eronder (een kop van twee delen).',
 'delivery.js › mailGeleverd() — de zin staat los in die functie en niet onder een sleutel.',
 'PricingPage.astro › in de opmaak — hij staat rechtstreeks in de opmaak; zoeken op de zin zelf vindt hem.',
 'Een tilde ~ erachter betekent dat het adres op een DEEL van de zin is gevonden: de zin wordt samengesteld '
 'en staat daar niet letterlijk zo.',
])

zet(f'A{rij}', 'Zo ziet een ingevulde rij eruit', vet=True, maat=12, kleur=GROEN); rij += 2

vb_kop = ['id', 'groep', 'waar', 'bron', 'Engels nu', 'Engels nieuw']
vb = ['home-4a1c8', 'Pagina', '/pricing/', 'PricingPage.astro › en.lead',
      'One price per product. It falls as you add.',
      'Clear rates with built-in volume discount.']
for k, (kop, waarde) in enumerate(zip(vb_kop, vb)):
    c = lm.cell(row=rij, column=k + 1, value=kop)
    c.font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor=KOPBALK)
    c.alignment = Alignment(wrap_text=True, vertical='center')
    d = lm.cell(row=rij + 1, column=k + 1, value=waarde)
    d.font = Font(name='Arial', size=10, color=INKT)
    d.alignment = Alignment(wrap_text=True, vertical='top')
    if kop.endswith('nieuw'):
        d.fill = PatternFill('solid', fgColor=INVUL)
for k in range(len(vb_kop)):
    lm.column_dimensions[get_column_letter(k + 1)].width = [22, 104, 20, 34, 40, 40][k] if k < 2 else [20, 34, 40, 40][k - 2]
lm.row_dimensions[rij + 1].height = 34
rij += 3
zet(f'A{rij}', 'Alleen de gele kolommen vul je in.', cursief=True, maat=10, kleur=LETOP)

# ══ blad 2 · Tekst ══════════════════════════════════════════════════════════
ws = wb.create_sheet('Tekst')
ws.sheet_properties.tabColor = KOPBALK

KOLOMMEN = [
    ('id', 13), ('groep', 17), ('waar', 24), ('blok', 26), ('soort', 11),
    ('bron', 34), ('let op', 34), ('Engels nu', 58), ('Nederlands nu', 58),
    ('Engels nieuw', 44), ('Nederlands nieuw', 44),
]
for i, (naam, breedte) in enumerate(KOLOMMEN, start=1):
    c = ws.cell(row=1, column=i, value=naam)
    c.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor=KOPBALK)
    c.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')
    ws.column_dimensions[get_column_letter(i)].width = breedte
ws.row_dimensions[1].height = 26

grens = Border(top=Side(style='thin', color=LIJN))
invul_vul = PatternFill('solid', fgColor=INVUL)
licht_vul = PatternFill('solid', fgColor=LICHT)

vorige_waar = None
for n, r in enumerate(rijen, start=2):
    letop = []
    if r['interp']:
        # gescheiden met een middenpunt en niet met een komma: in
        # `ex(ladderRate('complete', ENTRY_RUNG_LAST))` staat zelf al een komma,
        # en dan is niet meer te zien waar de ene interpolatie ophoudt.
        letop.append('de code vult in:  ' + '  ·  '.join(r['interp']))
    if r.get('waarschuwing'):
        letop.append('LET OP — ' + r['waarschuwing'])
    if r['elders']:
        letop.append(f"staat ook op {len(r['elders'])} andere pagina" + ('' if len(r['elders']) == 1 else "'s"))
    waarden = [r['id'], r['groep'], r['waar'], r['blok'], r['soort'], r['bron'],
               ' · '.join(letop), r['en'] or '', r['nl'] or '', '', '']
    nieuwe_groep = r['waar'] != vorige_waar
    vorige_waar = r['waar']
    for k, v in enumerate(waarden, start=1):
        c = ws.cell(row=n, column=k, value=v)
        c.font = Font(name='Arial', size=10,
                      color=LETOP if k == 7 and v else ('6B6D5C' if k in (1, 4, 5, 6) else INKT))
        c.alignment = Alignment(wrap_text=k >= 4, vertical='top')
        if k in (10, 11):
            c.fill = invul_vul
        elif k <= 3:
            c.fill = licht_vul
        if nieuwe_groep:
            c.border = grens

ws.freeze_panes = 'D2'
ws.auto_filter.ref = f'A1:K{len(rijen) + 1}'

import sys
pad = sys.argv[1] if len(sys.argv) > 1 else 'tekstronde/VISUAILS-tekstronde.xlsx'
wb.save(pad)
print('geschreven:', pad, '·', len(rijen), 'regels')
